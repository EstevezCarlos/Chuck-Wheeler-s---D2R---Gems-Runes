# import pandas as pd
import logging
# import functools as ft
import openpyxl
import json
# import os

logging.basicConfig(level=logging.INFO)

ORIGINAL = {
    "gems": 'files/original/data/global/excel/gems.txt',
    "misc": 'files/original/data/global/excel/misc.txt',
}
OUTPUT = {
    "gems": 'C:/Program Files (x86)/Diablo II Resurrected/Mods/ChuckCartwright/ChuckCartwright.mpq/data/global/excel/gems.txt',
    "misc": 'C:/Program Files (x86)/Diablo II Resurrected/Mods/ChuckCartwright/ChuckCartwright.mpq/data/global/excel/misc.txt',
    # "gems": 'ChuckCartwright.MPQ/data/globa/excel/gems.txt',
}

MODFILE_PATH = "C:/Program Files (x86)/Diablo II Resurrected/Mods/ChuckCartwright/ChuckCartwright.mpq/modinfo.json"
MODFILE_DICT = {
    "name": "ChuckCartwright",
    "savepath": "ChuckCartwright/",
}


def main():
    logging.info('main')
    Mod.absorb_files('files/source/gems.xlsx', 'gems')
    Mod.absorb_files('files/source/gems.xlsx', 'misc')
    Mod.write_files()
    with open(MODFILE_PATH, "w") as modfile:
        json.dump(MODFILE_DICT, modfile)


class Transformer:

    @classmethod
    def excel_to_matrix(cls, path):
        new_data_sheet = openpyxl.load_workbook(path).worksheets[0]
        return [[cell.value if cell.value != None else "" for cell in row] for row in new_data_sheet.iter_rows()]
    
    @classmethod
    def matrix_to_dicts(cls, matrix):
        head = matrix[0]
        data = matrix[1:]
        dicts =  {row[0]: {head[i]: val for i, val in enumerate(row)}for row in data if row[0]}
        # logging.info(dicts)
        return dicts

    @classmethod
    def excel_to_dicts(cls, path):

        # def excel_to_matrix(path):
        #     new_data_sheet = openpyxl.load_workbook(path).worksheets[0]
        #     return [[cell.value if cell.value != None else "" for cell in row] for row in new_data_sheet.iter_rows()]

        matrix = cls.excel_to_matrix(path)
        dicts = cls.matrix_to_dicts(matrix)
        return dicts

    @classmethod
    def csv_to_dicts(cls, path):
        with open(path, 'r') as file:
            matrix = [row.replace('\n','').split('\t') for row in file.readlines()]
        dicts = cls.matrix_to_dicts(matrix)
        return dicts


    @classmethod
    def dicts_to_tsv_string(cls, data):
        rows = list(data.values())
        string = ""

        for row in rows[:1]:
            column_names = list(row.keys())
            string += f'{cls.list_to_tab_seperated_string(column_names)}\n'

        for row in rows:
            row_values = list(row.values())
            string += f'{cls.list_to_tab_seperated_string(row_values)}\n'

        return string
        # with open(path, 'w') as f:
                # print(*za_listo, sep='\t')
                # print(*za_listo, sep='\t', file=f)
    @staticmethod
    def list_to_tab_seperated_string(data):

        string = ""

        for val in data[:1]:
            string += f'{val}'

        for val in data[1:]:
            string += f'\t{val}'

        return string

    @staticmethod
    def merge_dicts(modified_dicts, new_data):
        for i in new_data:
            for j in new_data[i]:
                if j in modified_dicts[i]:
                    modified_dicts[i][j] = new_data[i][j]
        return modified_dicts
        
class Mod:
    output_data = dict()

    @classmethod
    def absorb_files(cls, source_path, sheet):
        print(f'Absorbing {source_path}::{sheet}')
        logging.debug(f'absorb_files({cls}, {source_path}, {sheet})')
        new_data = Transformer.excel_to_dicts(source_path)
        if sheet not in cls.output_data:
            cls.output_data[sheet] = Transformer.csv_to_dicts(ORIGINAL[sheet])
            logging.debug(f'{sheet=}{cls.output_data=}')
        Transformer.merge_dicts(cls.output_data[sheet], new_data)

        logging.debug(f"### merged_data ###\n{cls.output_data}")

    @classmethod
    def write_files(cls):
        for filename, data in cls.output_data.items():
            with open(OUTPUT[filename], 'w') as file:
                string = Transformer.dicts_to_tsv_string(data)
                # logging.info(string.replace('\t','\\t'))
                file.write(string)

if __name__ == '__main__':

    main()
